from flask import Flask, render_template, request, jsonify, send_file, session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import os
import json
from collections import defaultdict

app = Flask(__name__)
app.secret_key = 'rostering_secret_key_2025'

# Session configuration for production (Render)
app.config['SESSION_COOKIE_SECURE'] = True  # Required for HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = 3600  # 1 hour

DATA_FILE = 'roster_data.json'
EXCEL_FILE = 'roster.xlsx'

# Admin and User credentials
CREDENTIALS = {
    'admin': 'aliyan123',
    'user': 'user123'
}

# Base shift pattern: Day, Night, Off, Off (repeating)
BASE_PATTERN = ['D', 'N', 'O', 'O']

def initialize_data():
    """Initialize roster data if not exists"""
    if not os.path.exists(DATA_FILE):
        employees = [
            {"name": "Namita Chowdhry", "emp_id": "85005100"},
            {"name": "Santhosh Rajaram Bathula(A320)", "emp_id": "85009008"},
            {"name": "Kaustubh B(B737)", "emp_id": "85002606"},
            {"name": "Vijay Gupta (B737)", "emp_id": "85002536"},
            {"name": "Manu Mohan (A320&B737)", "emp_id": "85005688"},
            {"name": "Macken Vaz(Tr)", "emp_id": "85003846"},
            {"name": "Karthikeyan(A320)", "emp_id": "85008568"},
            {"name": "Mahadev S. (A320)", "emp_id": "85005686"},
            {"name": "Vivek Chauhan (B737)", "emp_id": "85005699"},
            {"name": "Balamurugan Alagarsamy(A320)", "emp_id": "85008436"},
            {"name": "Jeet Pati (737)", "emp_id": "85011116"},
            {"name": "Siddique(Tr)", "emp_id": "85002426"},
            {"name": "Anal Kumar(Tr)", "emp_id": "85005698"},
            {"name": "Pravin C(A320)", "emp_id": "85002324"},
            {"name": "Dinesh S (8737)", "emp_id": "85003309"},
            {"name": "Maniraj S (A320)", "emp_id": "85008569"},
            {"name": "Ashish KS (A320&B737)", "emp_id": "85005689"},
            {"name": "Praful K (B737)", "emp_id": "85005690"},
            {"name": "Vivian Reginald Philip Pais (A320)", "emp_id": "85009317"},
            {"name": "Jefferson(A320)", "emp_id": "85008549"},
            {"name": "Abhimanyu (320)", "emp_id": "85007148"},
            {"name": "Vishnu Deva (A320)", "emp_id": "85008996"},
            {"name": "Janarthanan S(320)", "emp_id": "85005687"},
            {"name": "Ashish Kumar Verma(B737)", "emp_id": "85002951"},
            {"name": "Pravin J", "emp_id": "85002325"},
            {"name": "Vivek Pataskar(A320)", "emp_id": "85004114"},
            {"name": "Kshiteej S (8737)", "emp_id": "85002300"},
            {"name": "Prem (A320)", "emp_id": "85008783"},
            {"name": "Rashu(A320)", "emp_id": "85009429"},
            {"name": "Chetan kumar Gururani (A320)", "emp_id": "85008265"},
            {"name": "Yemuna BL", "emp_id": "85005685"},
            {"name": "Narender (A320)", "emp_id": "85004116"},
            {"name": "Sagar (A320)", "emp_id": "85008835"},
            {"name": "Soumee", "emp_id": "85005088"},
            {"name": "Nikita J(B737)", "emp_id": "85002171"}
        ]
        
        # Initialize roster for current month
        today = datetime.now()
        year = today.year
        month = today.month
        days_in_month = 31
        
        roster = {}
        
        # Generate roster with base pattern D-N-O-O, staggered for each employee
        for idx, emp in enumerate(employees):
            emp_roster = []
            # Stagger start position for each employee
            start_offset = idx % len(BASE_PATTERN)
            
            for day in range(days_in_month):
                pattern_idx = (start_offset + day) % len(BASE_PATTERN)
                shift = BASE_PATTERN[pattern_idx]
                emp_roster.append(shift)
            
            roster[emp['emp_id']] = emp_roster
        
        data = {
            'employees': employees,
            'roster': roster,
            'leave_applications': [],
            'year': year,
            'month': month,
            'leave_history': {},
            'pulled_staff': {},  # Track who was pulled and when
            'compensatory_offs': {}  # Track comp offs owed
        }
        
        save_data(data)
        return data
    else:
        return load_data()

def load_data():
    """Load data from JSON file"""
    with open(DATA_FILE, 'r') as f:
        return json.load(f)

def save_data(data):
    """Save data to JSON file"""
    with open(DATA_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def calculate_leave_priority(emp_id, data, applied_date):
    """Calculate leave priority based on history and application time"""
    history = data.get('leave_history', {}).get(emp_id, [])
    
    # Count leaves in last 3 months
    recent_leaves = 0
    cutoff_date = datetime.now() - timedelta(days=90)
    
    for leave in history:
        leave_date = datetime.strptime(leave['from'], '%Y-%m-%d')
        if leave_date >= cutoff_date:
            recent_leaves += 1
    
    # Base priority: 100 minus leaves taken (fewer leaves = higher priority)
    base_priority = 100 - (recent_leaves * 10)
    
    # Add time priority: earlier applications get bonus
    applied_dt = datetime.strptime(applied_date, '%Y-%m-%d %H:%M:%S')
    days_since_applied = (datetime.now() - applied_dt).days
    time_bonus = min(days_since_applied * 2, 20)  # Max 20 bonus points
    
    total_priority = max(base_priority + time_bonus, 0)
    return total_priority

def get_shift_type(shift_code):
    """Determine shift type from code"""
    if shift_code in ['D', 'D (OT)']:
        return 'day'
    elif shift_code in ['N', 'N (OT)']:
        return 'night'
    elif shift_code in ['O', 'C/O']:
        return 'off'
    elif shift_code in ['P/L', 'S/L', 'E/L', 'C/L']:
        return 'leave'
    else:
        return 'other'

def can_be_pulled(emp_id, date_idx, data):
    """Check if employee can be pulled for duty on this date"""
    roster = data['roster']
    
    # Check if already on duty
    current_shift = roster[emp_id][date_idx]
    if get_shift_type(current_shift) in ['day', 'night']:
        return False
    
    # Check previous day shift
    if date_idx > 0:
        prev_shift = roster[emp_id][date_idx - 1]
        prev_type = get_shift_type(prev_shift)
        
        # Cannot pull if they had night shift previous day
        if prev_type == 'night':
            return False
        
        # Can pull if they had day shift or off
        if prev_type in ['day', 'off']:
            return True
    
    return False

def find_replacement(emp_id, date_idx, required_shift_type, data):
    """Find replacement for emergency leave following proper rules"""
    roster = data['roster']
    employees = data['employees']
    
    # Get current shift counts
    day_count, night_count = count_shifts_by_type(date_idx, data)
    
    available = []
    for emp in employees:
        e_id = emp['emp_id']
        if e_id == emp_id:
            continue
        
        if can_be_pulled(e_id, date_idx, data):
            current_shift = roster[e_id][date_idx]
            
            # Only consider if they're off
            if get_shift_type(current_shift) == 'off':
                # Check if this is their first or second off
                if date_idx > 0:
                    prev_shift = roster[e_id][date_idx - 1]
                    prev_type = get_shift_type(prev_shift)
                    
                    # Calculate priority based on leave history and current pattern
                    priority_score = calculate_pull_priority(e_id, date_idx, data)
                    available.append({
                        'emp_id': e_id,
                        'priority': priority_score,
                        'prev_type': prev_type
                    })
    
    # Sort by priority (higher is better to pull)
    available.sort(key=lambda x: x['priority'], reverse=True)
    
    return available[0]['emp_id'] if available else None

def calculate_pull_priority(emp_id, date_idx, data):
    """Calculate priority for pulling someone (lower is better to pull)"""
    pulled_history = data.get('pulled_staff', {}).get(emp_id, [])
    
    # Those pulled less recently have lower priority (better to pull)
    recent_pulls = len([p for p in pulled_history if p > date_idx - 7])
    
    # Factor in compensatory offs owed
    comp_offs_owed = data.get('compensatory_offs', {}).get(emp_id, 0)
    
    # Lower score = higher priority to be pulled
    priority = recent_pulls * 10 + comp_offs_owed * 5
    return priority

def update_pattern_after_pull(emp_id, date_idx, pull_type, data):
    """Update roster pattern after pulling someone
    pull_type: 'first_off' or 'second_off'
    """
    roster = data['roster']
    
    if pull_type == 'first_off':
        # Pattern becomes: D, N, N, O, D
        # They worked D, now N, pulled for first O -> becomes N
        roster[emp_id][date_idx] = 'N (OT)'
        
        # Next day should be O, then D
        if date_idx + 1 < len(roster[emp_id]):
            roster[emp_id][date_idx + 1] = 'O'
        if date_idx + 2 < len(roster[emp_id]):
            if get_shift_type(roster[emp_id][date_idx + 2]) == 'off':
                roster[emp_id][date_idx + 2] = 'D'
    
    elif pull_type == 'second_off':
        # Pattern becomes: D, N, O, D, D
        # They worked D, N, had O, pulled for second O -> becomes D
        roster[emp_id][date_idx] = 'D (OT)'
        
        # Next day should also be D
        if date_idx + 1 < len(roster[emp_id]):
            if get_shift_type(roster[emp_id][date_idx + 1]) == 'off':
                roster[emp_id][date_idx + 1] = 'D'
    
    # Track the pull
    if 'pulled_staff' not in data:
        data['pulled_staff'] = {}
    if emp_id not in data['pulled_staff']:
        data['pulled_staff'][emp_id] = []
    data['pulled_staff'][emp_id].append(date_idx)
    
    # Update compensatory off tracking
    if 'compensatory_offs' not in data:
        data['compensatory_offs'] = {}
    if emp_id not in data['compensatory_offs']:
        data['compensatory_offs'][emp_id] = 0
    data['compensatory_offs'][emp_id] += 1

def count_shifts_by_type(date_idx, data):
    """Count day and night shifts for a date"""
    roster = data['roster']
    day_count = 0
    night_count = 0
    
    for emp_id, shifts in roster.items():
        if date_idx < len(shifts):
            shift = shifts[date_idx]
            shift_type = get_shift_type(shift)
            if shift_type == 'day':
                day_count += 1
            elif shift_type == 'night':
                night_count += 1
    
    return day_count, night_count

def generate_excel(data):
    """Generate Excel file with roster"""
    wb = Workbook()
    
    # Sheet 1: Main Roster
    ws_roster = wb.active
    ws_roster.title = "Roster"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Days in month
    year = data['year']
    month = data['month']
    days_in_month = 31
    
    # Title row
    ws_roster.merge_cells('A1:B1')
    title_cell = ws_roster.cell(1, 1, f"STAFF ROSTER - {datetime(year, month, 1).strftime('%B %Y')}")
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Header row: Day numbers
    ws_roster.cell(2, 1, "NAME").fill = header_fill
    ws_roster.cell(2, 1).font = header_font
    ws_roster.cell(2, 1).border = border
    ws_roster.cell(2, 2, "EMP ID").fill = header_fill
    ws_roster.cell(2, 2).font = header_font
    ws_roster.cell(2, 2).border = border
    
    for day in range(1, days_in_month + 1):
        cell = ws_roster.cell(2, day + 2, day)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Header row: Day names
    ws_roster.cell(3, 1, "").fill = header_fill
    ws_roster.cell(3, 1).border = border
    ws_roster.cell(3, 2, "").fill = header_fill
    ws_roster.cell(3, 2).border = border
    
    day_names = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]
    for day in range(1, days_in_month + 1):
        day_name = day_names[(day - 1) % 7]
        cell = ws_roster.cell(3, day + 2, day_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Employee data
    employees = data['employees']
    roster = data['roster']
    
    for idx, emp in enumerate(employees):
        row = idx + 4
        ws_roster.cell(row, 1, emp['name']).border = border
        ws_roster.cell(row, 2, emp['emp_id']).border = border
        
        emp_roster = roster.get(emp['emp_id'], [])
        for day in range(1, min(days_in_month + 1, len(emp_roster) + 1)):
            shift = emp_roster[day - 1] if day - 1 < len(emp_roster) else ""
            cell = ws_roster.cell(row, day + 2, shift)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            # Color coding for shifts
            if 'D' in shift:
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif any(n in shift for n in ['N', 'N/O', 'N3']):
                cell.fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            elif shift == 'O':
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            elif 'P/L' in shift:
                cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
            elif 'S/L' in shift or 'E/L' in shift:
                cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            elif 'C/O' in shift:
                cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
            elif 'OT' in shift:
                cell.fill = PatternFill(start_color="FFE5B4", end_color="FFE5B4", fill_type="solid")
                cell.font = Font(bold=True, color="FF0000")
    
    # Adjust column widths
    ws_roster.column_dimensions['A'].width = 35
    ws_roster.column_dimensions['B'].width = 12
    from openpyxl.utils import get_column_letter
    for col in range(3, days_in_month + 3):
        ws_roster.column_dimensions[get_column_letter(col)].width = 6
    
    # Sheet 2: Leave Applications
    ws_leave = wb.create_sheet("Leave Applications")
    
    headers = ["S.No", "Employee Name", "Employee ID", "From Date", "To Date", "Leave Type", "Status", "Priority", "Applied Date", "Reason"]
    for col, header in enumerate(headers, 1):
        cell = ws_leave.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    leave_apps = data.get('leave_applications', [])
    for idx, leave in enumerate(leave_apps, 1):
        ws_leave.cell(idx + 1, 1, idx).border = border
        ws_leave.cell(idx + 1, 2, leave['emp_name']).border = border
        ws_leave.cell(idx + 1, 3, leave['emp_id']).border = border
        ws_leave.cell(idx + 1, 4, leave['from_date']).border = border
        ws_leave.cell(idx + 1, 5, leave['to_date']).border = border
        ws_leave.cell(idx + 1, 6, leave['leave_type']).border = border
        
        status_cell = ws_leave.cell(idx + 1, 7, leave['status'])
        status_cell.border = border
        if leave['status'] == 'Approved':
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100", bold=True)
        elif leave['status'] == 'Rejected':
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="9C0006", bold=True)
        else:
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            status_cell.font = Font(color="9C6500", bold=True)
        
        ws_leave.cell(idx + 1, 8, leave.get('priority', 0)).border = border
        ws_leave.cell(idx + 1, 9, leave.get('applied_date', '')).border = border
        ws_leave.cell(idx + 1, 10, leave.get('reason', '')).border = border
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws_leave.column_dimensions[col].width = 15
    ws_leave.column_dimensions['B'].width = 35
    ws_leave.column_dimensions['J'].width = 40
    
    # Sheet 3: Daily Statistics
    ws_stats = wb.create_sheet("Daily Statistics")
    
    stat_headers = ["Date", "Day", "Day Shift Count", "Night Shift Count", "On Leave", "Total Working"]
    for col, header in enumerate(stat_headers, 1):
        cell = ws_stats.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    for day in range(1, days_in_month + 1):
        day_count, night_count = count_shifts_by_type(day - 1, data)
        
        # Count leaves
        leave_count = 0
        for emp_id, shifts in roster.items():
            if day - 1 < len(shifts):
                if get_shift_type(shifts[day - 1]) == 'leave':
                    leave_count += 1
        
        date_obj = datetime(year, month, day)
        day_name = date_obj.strftime('%A')
        
        ws_stats.cell(day + 1, 1, date_obj.strftime('%Y-%m-%d')).border = border
        ws_stats.cell(day + 1, 2, day_name).border = border
        ws_stats.cell(day + 1, 3, day_count).border = border
        ws_stats.cell(day + 1, 4, night_count).border = border
        ws_stats.cell(day + 1, 5, leave_count).border = border
        ws_stats.cell(day + 1, 6, day_count + night_count).border = border
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_stats.column_dimensions[col].width = 18
    
    # Sheet 4: Overtime & Comp Off Tracking
    ws_ot = wb.create_sheet("Overtime Tracking")
    
    ot_headers = ["Employee Name", "Employee ID", "Times Pulled", "Comp Offs Owed", "Last Pulled Date"]
    for col, header in enumerate(ot_headers, 1):
        cell = ws_ot.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    pulled_staff = data.get('pulled_staff', {})
    comp_offs = data.get('compensatory_offs', {})
    
    row = 2
    for emp in employees:
        emp_id = emp['emp_id']
        if emp_id in pulled_staff and pulled_staff[emp_id]:
            times_pulled = len(pulled_staff[emp_id])
            comp_owed = comp_offs.get(emp_id, 0)
            last_pulled = max(pulled_staff[emp_id]) if pulled_staff[emp_id] else 0
            last_pulled_date = f"Day {last_pulled + 1}" if last_pulled else "Never"
            
            ws_ot.cell(row, 1, emp['name']).border = border
            ws_ot.cell(row, 2, emp_id).border = border
            ws_ot.cell(row, 3, times_pulled).border = border
            ws_ot.cell(row, 4, comp_owed).border = border
            ws_ot.cell(row, 5, last_pulled_date).border = border
            row += 1
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_ot.column_dimensions[col].width = 20
    ws_ot.column_dimensions['A'].width = 35
    
    # Sheet 5: Legend
    ws_legend = wb.create_sheet("Legend")
    
    legend_data = [
        ["Shift Code", "Description", "Type"],
        ["D", "Day Shift (Regular)", "Working"],
        ["N", "Night Shift (Regular)", "Working"],
        ["O", "Off Day", "Rest"],
        ["D (OT)", "Day Shift - Overtime (Pulled)", "Working + OT"],
        ["N (OT)", "Night Shift - Overtime (Pulled)", "Working + OT"],
        ["P/L", "Planned Leave (Pre-approved)", "Leave"],
        ["S/L", "Sick Leave (Emergency)", "Leave"],
        ["E/L", "Emergency Leave (Immediate)", "Leave"],
        ["C/L", "Casual Leave", "Leave"],
        ["C/O", "Compensatory Off (Earned)", "Rest"],
    ]
    
    for row_idx, row_data in enumerate(legend_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_legend.cell(row_idx, col_idx, value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
            cell.border = border
    
    ws_legend.column_dimensions['A'].width = 15
    ws_legend.column_dimensions['B'].width = 40
    ws_legend.column_dimensions['C'].width = 15
    
    wb.save(EXCEL_FILE)
    return EXCEL_FILE

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login', methods=['POST'])
def login():
    try:
        data = request.json
        username = data.get('username')
        password = data.get('password')
        
        if username in CREDENTIALS and CREDENTIALS[username] == password:
            session['user'] = username
            session['role'] = username
            return jsonify({'success': True, 'role': username})
        else:
            return jsonify({'success': False, 'message': 'Invalid credentials'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'})

@app.route('/logout', methods=['POST'])
def logout():
    session.pop('user', None)
    session.pop('role', None)
    return jsonify({'success': True})

@app.route('/check-auth')
def check_auth():
    if 'user' in session:
        return jsonify({'authenticated': True, 'role': session['role']})
    return jsonify({'authenticated': False})

@app.route('/get-roster')
def get_roster():
    data = load_data()
    return jsonify(data)

@app.route('/update-shift', methods=['POST'])
def update_shift():
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    emp_id = request.json.get('emp_id')
    date_idx = request.json.get('date_idx')
    new_shift = request.json.get('shift')
    
    data = load_data()
    if emp_id in data['roster']:
        data['roster'][emp_id][date_idx] = new_shift
        save_data(data)
        generate_excel(data)
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': 'Employee not found'})

@app.route('/apply-leave', methods=['POST'])
def apply_leave():
    if 'user' not in session:
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    data = load_data()
    leave_data = request.json
    
    applied_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Calculate priority
    priority = calculate_leave_priority(leave_data['emp_id'], data, applied_date)
    
    leave_app = {
        'id': len(data['leave_applications']) + 1,
        'emp_id': leave_data['emp_id'],
        'emp_name': leave_data['emp_name'],
        'from_date': leave_data['from_date'],
        'to_date': leave_data['to_date'],
        'leave_type': leave_data['leave_type'],
        'reason': leave_data.get('reason', ''),
        'status': 'Pending',
        'priority': priority,
        'applied_date': applied_date
    }
    
    data['leave_applications'].append(leave_app)
    save_data(data)
    generate_excel(data)
    
    return jsonify({'success': True, 'message': 'Leave application submitted', 'priority': priority})

@app.route('/approve-leave', methods=['POST'])
def approve_leave():
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    leave_id = request.json.get('leave_id')
    action = request.json.get('action')
    
    data = load_data()
    
    for leave in data['leave_applications']:
        if leave['id'] == leave_id:
            leave['status'] = 'Approved' if action == 'approve' else 'Rejected'
            
            if action == 'approve':
                # Update roster with leave
                emp_id = leave['emp_id']
                from_date = datetime.strptime(leave['from_date'], '%Y-%m-%d')
                to_date = datetime.strptime(leave['to_date'], '%Y-%m-%d')
                
                is_emergency = leave['leave_type'] in ['Sick Leave', 'Emergency Leave']
                
                current_date = from_date
                while current_date <= to_date:
                    day_idx = current_date.day - 1
                    if day_idx < len(data['roster'][emp_id]):
                        original_shift = data['roster'][emp_id][day_idx]
                        
                        # Mark as leave based on type
                        if leave['leave_type'] == 'Sick Leave':
                            data['roster'][emp_id][day_idx] = 'S/L'
                        elif leave['leave_type'] == 'Emergency Leave':
                            data['roster'][emp_id][day_idx] = 'E/L'
                        elif leave['leave_type'] == 'Casual Leave':
                            data['roster'][emp_id][day_idx] = 'C/L'
                        else:
                            data['roster'][emp_id][day_idx] = 'P/L'
                        
                        # Find replacement if emergency and they were on duty
                        if is_emergency and get_shift_type(original_shift) in ['day', 'night']:
                            replacement = find_replacement(emp_id, day_idx, get_shift_type(original_shift), data)
                            if replacement:
                                # Determine if first or second off
                                if day_idx > 0:
                                    prev_shift = data['roster'][replacement][day_idx - 1]
                                    if get_shift_type(prev_shift) == 'night':
                                        # First off after night shift
                                        update_pattern_after_pull(replacement, day_idx, 'first_off', data)
                                    else:
                                        # Second off
                                        update_pattern_after_pull(replacement, day_idx, 'second_off', data)
                    
                    current_date += timedelta(days=1)
                
                # Update leave history
                if 'leave_history' not in data:
                    data['leave_history'] = {}
                if emp_id not in data['leave_history']:
                    data['leave_history'][emp_id] = []
                data['leave_history'][emp_id].append({
                    'from': leave['from_date'],
                    'to': leave['to_date'],
                    'type': leave['leave_type']
                })
            
            save_data(data)
            generate_excel(data)
            return jsonify({'success': True})
    
    return jsonify({'success': False, 'message': 'Leave application not found'})

@app.route('/download-excel')
def download_excel():
    data = load_data()
    generate_excel(data)
    return send_file(EXCEL_FILE, as_attachment=True, download_name=f'roster_{data["month"]}_{data["year"]}.xlsx')

if __name__ == '__main__':
    initialize_data()
    data = load_data()
    generate_excel(data)
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('DEBUG', 'False') == 'True'
    app.run(host='0.0.0.0', port=port, debug=debug)
